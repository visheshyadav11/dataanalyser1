from email import message
from fileinput import filename
from pstats import Stats
from turtle import clear, color
from flask import Flask, render_template, request
import os
import pandas as pd
import numpy as np
from tqdm.notebook import tqdm_notebook
from openpyxl.styles import Alignment
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import openpyxl
import matplotlib.pyplot as plt
import datetime
import files
import requests

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    global file_path,mapping_file_path,file,mapping_file,colorNTPR, colorMAT05, colorMAT10, colorMAT20, colorMAT30, portfolio_type, status_budget_tls,start_year,end_year,yearrange,color_range
    if request.method == 'POST':

        file = request.files['file']
        file_path = file.filename
        file.save(file_path)
        
        # mapping_file="https://stmicroelectronics.sharepoint.com/:x:/r/teams/TDP_TPPM_PMO/_layouts/15/Doc.aspx?sourcedoc=%7BD33063BA-D9C2-46B8-B28B-193DB18E888A%7D&file=Program%20mapping_V1.5.1.xlsx&action=default&mobileredirect=true"
        

        mapping_file_url = 'https://stmicroelectronics.sharepoint.com/:x:/r/teams/TDP_TPPM_PMO/_layouts/15/Doc.aspx?sourcedoc=%7BD33063BA-D9C2-46B8-B28B-193DB18E888A%7D&file=Program%20mapping_V1.5.1.xlsx&action=default&mobileredirect=true'
        mapping_file_path = 'mapping_file.xlsx'

        response = requests.get(mapping_file_url)

        with open(mapping_file_path, 'wb') as f:
            f.write(response.content)
        
        
        # mapping_file = request.files['mapping_file']
        # mapping_file_path = mapping_file.filename
        # mapping_file.save(mapping_file_path)

        start_year = request.form['start_year']
        end_year = request.form['end_year']
        yearrange=str(start_year)+'-'+str(end_year)
        colorNTPR = request.form['colorNTPR']
        colorMAT05 = request.form['colorMAT05']
        colorMAT10 = request.form['colorMAT10']
        colorMAT20 = request.form['colorMAT20']
        colorMAT30 = request.form['colorMAT30']
        portfolio_type = request.form.getlist('portfolio_type[]')
        portfolio_type=', '.join(portfolio_type)
        colorNTPR=colorNTPR[1:]
        colorMAT05=colorMAT05[1:]
        colorMAT10=colorMAT10[1:]
        colorMAT20=colorMAT20[1:]
        colorMAT30=colorMAT30[1:]

        color_range = {'NTPR' : (colorNTPR),
               'MAT05' : (colorMAT05),
               'MAT10' : (colorMAT10),
               'MAT20' : (colorMAT20),
               'MAT30' : (colorMAT30)}

        status_budget_tls = request.form.getlist('status_budget_tls[]')
        status_budget_tls=', '.join(status_budget_tls)
       
        data_analysis = DataAnalysis(file_path, color_range, yearrange, portfolio_type, status_budget_tls, sorting_parameter, ascending)
        data_analysis.load_dataset(file_path, mapping_file_path)

        message = 'File uploaded successfully!'+f"You selected the year range {yearrange},the portfolio type is {(portfolio_type)}, and the status budget TLS is {(status_budget_tls)}"
        return render_template('index.html', message=message)
    return render_template('index.html')

class DataAnalysis():
    # global file_path, portfolio_type, status_budget_tls,start_year,end_year,yearrange,ascending,sorting_parameter
    global file_path,mapping_file_path,colorNTPR, colorMAT05, colorMAT10, colorMAT20, colorMAT30, portfolio_type, status_budget_tls,start_year,end_year,yearrange,color_range,sorting_parameter,ascending

    sorting_parameter = 'Portfolio, Program, SubProgram'
    ascending = True
    
    def __init__(self, filename,color_schema, yearrangefactor, portfolio_type, status_budget_tls, sorting_parameter, ascending = True):
        self._filename = filename
        self._input_sheet = 'Roadmap'
        self._mappingsheet = 'Mapping'
        self._colorschemadict = color_schema
        self._quarterinformation = {1 : 1, 2 : 4, 3 : 7, 4 : 10}
        self._writingoutputfile = 'output.xlsx'
        self._yearlist = []
        self._startingcolumn = 'E'
        self._yearheader = 1
        self._quarterheader = 2
        self._yearnameheader = 3
        self.portfolio_type = portfolio_type
        self.status_budget_tls = status_budget_tls
        self.yearrangefactor = yearrangefactor
        self.sorting_parameter = sorting_parameter
        self.ascending = ascending
        self._writingmonths = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'June', 'July', 'Aug', 'Sept', 'Oct', 'Nov', 'Dec']

    def load_dataset(self, file, mapping_file):
        if file.endswith('xlsx') or file.endswith('xlsm') and mapping_file.endswith('xlsx'):
            self._inputdata = pd.read_excel(file, sheet_name = self._input_sheet)
            self._inputdata = self._inputdata.rename(columns = lambda x: x.strip())
            self._mappingsheet = pd.read_excel(mapping_file, sheet_name = self._mappingsheet)
            self._mappingsheet = self._mappingsheet.rename(columns = lambda x: x.strip())
            self._settingstatusbudget()
            self._sortingdataframe()
        else:
            return 'Invalid File Format uploaded'

    def _settingstatusbudget(self):
        if str(self.status_budget_tls) != 'None':
            _tempstatusdataframe = pd.DataFrame()
            for _statusbudgettls in self.status_budget_tls.split(','):
                _tempstatusdataframe = pd.concat([_tempstatusdataframe, self._inputdata[self._inputdata['Status to use for Budget / TLS'] == _statusbudgettls.strip()]])
            self._inputdata = _tempstatusdataframe

    def _sortingdataframe(self):
        self._inputdata = self._inputdata[["IFRS Code (TPP PLM Marketing Name)", "NTPR", "MAT05", "MAT10", "MAT20", "MAT30"]]
        self._inputdata['Portfolio'] = ''
        self._inputdata['Program'] = ''
        self._inputdata['SubProgram'] = ''
        for _row in range(self._inputdata.shape[0]):
            _taskname = self._inputdata['IFRS Code (TPP PLM Marketing Name)'].iloc[_row]
            _tempmapping = self._mappingsheet[self._mappingsheet['title'] == _taskname]
            if _tempmapping.shape[0]:
                self._inputdata['Portfolio'].iloc[_row] = _tempmapping['PRG GRP1 (portfolio)'].iloc[0]
                self._inputdata['Program'].iloc[_row] = _tempmapping['PROGRAM'].iloc[0]
                self._inputdata['SubProgram'].iloc[_row] = _tempmapping['SUB PROGRAM'].iloc[0]
            else:
              print(f"ERROR ! For {_taskname} it couldn't be found in the Mapping Sheet")

        self._inputdata.rename(columns = {"IFRS Code (TPP PLM Marketing Name)" : "Technology"}, inplace = True)

        if str(self.sorting_parameter) != 'None':
            self.sorting_parameter = self.sorting_parameter.split(",")
            self.sorting_parameter = [x.strip() for x in self.sorting_parameter]
            self._inputdata.sort_values(by = self.sorting_parameter, ascending = self.ascending, inplace = True)

    def grant_chart_making(self):
        self._finalplotresult = list()
        _counter = 0
        flag = False
        for row in tqdm_notebook(range(self._inputdata.shape[0])):
            _taskname = self._inputdata.iloc[row, :].iloc[0]
            _ntpr = self._inputdata.iloc[row, :].loc["NTPR"]
            _mat05 = self._inputdata.iloc[row, :].loc["MAT05"]
            _mat10 = self._inputdata.iloc[row, :].loc["MAT10"]
            _mat20 = self._inputdata.iloc[row, :].loc["MAT20"]
            _mat30 = self._inputdata.iloc[row, :].loc["MAT30"]
            _tempresult = self._creatematplotlibplot(_ntpr, _mat05, _mat10, _mat20, _mat30, _taskname)
            if type(_tempresult) == list:
                self._finalplotresult.append(_tempresult)

    def _getmonthfromweek(self, date):
        if type(date) == float:
            return ''
        if "'" in date:
            week, year = date.split("'")
        elif "’" in date:
            week, year = date.split("’")
        year = str(20) + year
        try:
            self._yearlist.append(int(year))
        except ValueError:
            pass
        date = year + '-' + week
        date = date.replace('\u200b', '').replace('\xa0', '')
        return str(datetime.datetime.strptime(date + '-1', "%Y-W%W-%w")).split(' ')[0]

    def _getenddatantpr(self, date):
        year, month, day = int(date.split('-')[0]), int(date.split('-')[1]), int(date.split('-')[2])
        if month == 12:
            year = year + 1
            month = 1
        else:
            month = month + 1
        return str(year) + '-' + str(month) + '-' + str(day)

    def _getquarterstartingdate(self, date):
        year, month, day = int(date.split('-')[0]), int(date.split('-')[1]), int(date.split('-')[2])
        month = self._quarterinformation.get(pd.Timestamp(datetime.date(year, month, day)).quarter)
        return str(year) + '-' + str(month) + '-' + str(1)

    def _creatematplotlibplot(self, NTPR, MAT05, MAT10, MAT20, MAT30, task_name):
        if NTPR and str(NTPR)[0] == 'W':NTPR = self._getmonthfromweek(NTPR.replace(' ', ''))
        if MAT05 and str(MAT05)[0] == 'W':MAT05 = self._getmonthfromweek(MAT05.replace(' ', ''))
        if MAT10 and str(MAT10)[0] == 'W':MAT10 = self._getmonthfromweek(MAT10.replace(' ', ''))
        if MAT20 and str(MAT20)[0] == 'W':MAT20 = self._getmonthfromweek(MAT20.replace(' ', ''))
        if MAT30 and str(MAT30)[0] == 'W':MAT30 = self._getmonthfromweek(MAT30.replace(' ', ''))

        _reversedic = {NTPR : 'NTPR', MAT05 : 'MAT05', MAT10 : 'MAT10', MAT20 : 'MAT20', MAT30 : 'MAT30'}

        matlist = [NTPR, MAT05, MAT10, MAT20, MAT30]
        matlist = [x for x in matlist if type(x)!=float]
        matlist = [x for x in matlist if str(x)[0] == '2']

        _illegalflag = False

        if len(matlist) > 1:
            for _value in range(len(matlist) - 1):
                _currentdate = matlist[_value]
                _nextdate = matlist[_value + 1]
                _currentdateyear, _currentdatemonth, _currentdateday = int(_currentdate.split('-')[0]), int(_currentdate.split('-')[1]), int(_currentdate.split('-')[2])
                _nextdateyear, _nextdatemonth, _nextdateday = int(_nextdate.split('-')[0]), int(_nextdate.split('-')[1]), int(_nextdate.split('-')[2])
                _currentdate = datetime.datetime(_currentdateyear, _currentdatemonth, _currentdateday)
                _nextdate = datetime.datetime(_nextdateyear, _nextdatemonth, _nextdateday)
                if _currentdate > _nextdate:
                    print(f'Illegal date in Task {task_name} between Deadline {_reversedic.get(matlist[_value])} and {_reversedic.get(matlist[_value + 1])}')
                    _illegalflag = True

        if _illegalflag == True:
            return 'Illegal Dates no graph can be drawn'

        if not NTPR and not MAT05 and not MAT10 and not MAT20 and not MAT30:
            return 'NO Graph can be drawn'
        else:
            _tempdflist = list()
            flag = False
            processed = list()
            for mat in range(len(matlist)):
                if _reversedic.get(matlist[0]) == 'NTPR':
                    if flag == False:
                        _tempdf = dict(Task = task_name, Start = matlist[0], Finish = self._getenddatantpr(matlist[0]), Resource = _reversedic.get(str(matlist[mat])))
                        _previouscompletiondate = self._getenddatantpr(matlist[0])
                        flag = True
                    elif flag == True:
                        _tempdf = dict(Task = task_name, Start = _previouscompletiondate , Finish = matlist[mat], Resource = _reversedic.get(str(matlist[mat])))
                        _previouscompletiondate = self._getenddatantpr(matlist[mat])
                else:
                    if flag == False:
                        _tempdf = dict(Task = task_name, Start = self._getquarterstartingdate(matlist[mat]), Finish = matlist[mat], Resource = _reversedic.get(str(matlist[mat])))
                        flag = True
                        _previouscompletiondate = self._getenddatantpr(matlist[mat])
                    elif flag == True:
                        _tempdf = dict(Task = task_name, Start = _previouscompletiondate, Finish = matlist[mat], Resource = _reversedic.get(str(matlist[mat])))
                        _previouscompletiondate = self._getenddatantpr(matlist[mat])

                if _reversedic.get(str(matlist[mat])) not in processed:
                    _tempdflist.append(_tempdf)
                    processed.append(_reversedic.get(str(matlist[mat])))

            return _tempdflist

    def _findcellvalue(self, number):
        _startingalphabetnumber = ord(self._startingcolumn) - 65 + 1
        number = number + _startingalphabetnumber - 1
        if number > 26:
            _tempnumber = (number - 1) // 26
            _remainder = number % 26
            if _remainder == 0:
                _remainder = 26
            _tempalphabet = chr(_tempnumber + 65 - 1) + chr(_remainder + 65 - 1)
        else:
            _tempalphabet = chr(number + 65 - 1)
        return _tempalphabet

    def as_text(self, value):
        if value is None:
            return ""
        return str(value)

    def writingdatatoexcel(self):
        self._yearlist = list(set(self._yearlist))
        self._yearlist = sorted(self._yearlist)
        _yearletterdict = {}
        _yearstartdict = {}
        _yearenddict = {}
        wb = Workbook()
        sheet = wb.active
        _startvalue = 1
        _endvalue = 12
        _writingcounter = 5
        _quarterwritingcounter = 5
        for _year in self._yearlist:
            _start, _end = self._findcellvalue(_startvalue), self._findcellvalue(_endvalue)
            _yearstartdict[_year] = _start
            _yearenddict[_year] = _end
            _writingyearname = []
            for x in range(_startvalue, _endvalue + 1):
                _yearletterdict[self._findcellvalue(x)] = x
                _writingyearname.append(self._findcellvalue(x))
            _startvalue = _endvalue + 1
            _endvalue = _endvalue + 12
            sheet.merge_cells(f'{_start}{self._yearheader}:{_end}{self._yearheader}')
            cell = sheet.cell(row = self._yearheader, column = _writingcounter)
            cell.value = _year
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
            _writingcounter = _writingcounter + 12
            for _iteration in range(len(_writingyearname)):
                sheet[f'{_writingyearname[_iteration]}{self._yearnameheader}'] = self._writingmonths[_iteration]
                sheet[f'{_writingyearname[_iteration]}{self._yearnameheader}'].alignment = Alignment(textRotation = 90)
            for index, _iteration in enumerate(range(0, len(_writingyearname), 3)):
                _yearheaderstart, _yearheaderend = _writingyearname[_iteration:_iteration + 3][0], _writingyearname[_iteration:_iteration + 3][-1]
                sheet.merge_cells(f'{_yearheaderstart}{self._quarterheader}:{_yearheaderend}{self._quarterheader}')
                cell = sheet.cell(row = self._quarterheader, column = _quarterwritingcounter)
                _quarterwritingcounter = _quarterwritingcounter + 3
                cell.value = f'Q{index + 1}'
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center')

        sheet['A3'] = 'Portfolio'
        sheet['B3'] = 'PROGRAM'
        sheet['C3'] = 'SUB PROGRAM'
        sheet['D3'] = 'Technology'

        _yearletterdictreverse = {v : k for k, v in _yearletterdict.items()}

        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value

        _startingbgwritingcounter = 4
        _startingtempbgwritingcounter = 4
        _bgwritingcounter = 4
        _tempindex = list()
        _tempindexemptyyear = list()
        index = 0
        _writingindex = 0

        for _iteration in self._finalplotresult:
            if len(_iteration):
                # finding those rows where there is no data when year filter is applied
                _tempyearliststart = list(set([int(x['Start'].split('-')[0]) for x in _iteration]))
                _tempyearlistfinish = list(set([int(x['Finish'].split('-')[0]) for x in _iteration]))
                _tempyearlist = list(set(_tempyearliststart + _tempyearlistfinish))
                if str(self.yearrangefactor) != 'None':
                    _selectedyears = list()
                    for _yearselector in self.yearrangefactor.split(','):
                        if '-' in _yearselector:
                            _yearselector = _yearselector.split('-')
                            _yearselectorstart, _yearselectorend = int(_yearselector[0]), int(_yearselector[1])
                            _selectedyears.extend(list(np.arange(_yearselectorstart, _yearselectorend + 1)))
                        else:
                            _yearselector = int(_yearselector)
                            _selectedyears.append(_yearselector)
                    _selectedyears = sorted(_selectedyears)
                    _intersectionyears = [value for value in _tempyearlist if value in _selectedyears]
                else:
                    _intersectionyears = []

                if len(_intersectionyears) or str(self.yearrangefactor) == 'None':
                    index = index + 1
                    for _task in _iteration:
                        if len(_task):
                            _startdate, _enddate, _resource = _task['Start'], _task['Finish'], _task['Resource']
                            _startyear, _startmonth = int(_startdate.split('-')[0]), int(_startdate.split('-')[1])
                            _endyear, _endmonth = int(_enddate.split('-')[0]), int(_enddate.split('-')[1])
                            _startyear, _endyear = _yearstartdict.get(_startyear), _yearstartdict.get(_endyear)
                            _startmonth = _yearletterdict.get(_startyear) + _startmonth - 1
                            _endmonth = _yearletterdict.get(_endyear) + _endmonth - 1
                            for _row in range(_startmonth, _endmonth + 1):
                                _columnname = _yearletterdictreverse.get(_row)
                                if _resource == 'NTPR':
                                    sheet[f'{_columnname}{_bgwritingcounter}'].fill = PatternFill(start_color = self._colorschemadict.get(_resource), fill_type = "solid")
                                    break
                                else:
                                    sheet[f'{_columnname}{_bgwritingcounter}'].fill = PatternFill(start_color = self._colorschemadict.get(_resource), fill_type = "solid")

                    sheet[f'D{_bgwritingcounter}'] = _iteration[0]['Task']
                    _tempmapping = self._mappingsheet[self._mappingsheet['title'] == _iteration[0]['Task']]
                    if _tempmapping.shape[0]:
                        for _portfolio in self.portfolio_type.split(','):
                            if _portfolio.lower().strip() in _tempmapping['PRG GRP1 (portfolio)'].iloc[0].lower().strip():
                                _tempindex.append(index + _startingbgwritingcounter - 1)
                        sheet[f'A{_bgwritingcounter}'] = _tempmapping['PRG GRP1 (portfolio)'].iloc[0]
                        sheet[f'B{_bgwritingcounter}'] = _tempmapping['PROGRAM'].iloc[0]
                        sheet[f'C{_bgwritingcounter}'] = _tempmapping['SUB PROGRAM'].iloc[0]
                    _bgwritingcounter = _bgwritingcounter + 1

                _writingindex = _writingindex + 1

        _tempiteratorcounter = 0

        if str(portfolio_type) != 'None':
            for _iterator in range(_startingbgwritingcounter, _bgwritingcounter + 1):
                if _iterator not in _tempindex:
                    _tempiteratorcounter = _tempiteratorcounter + 1
                else:
                    if _tempiteratorcounter == 0:
                        _startingbgwritingcounter = _startingbgwritingcounter + 1
                        _tempiteratorcounter = 0
                    else:
                        sheet.delete_rows(_startingbgwritingcounter, _tempiteratorcounter)
                        _startingbgwritingcounter = _startingbgwritingcounter + 1
                        _tempiteratorcounter = 0

            sheet.delete_rows(_startingbgwritingcounter, _tempiteratorcounter)

        _selectedyears = list()
        if str(self.yearrangefactor) != 'None':
            _columnstartingcounter = ord(self._startingcolumn) - 65 + 1 - 1
            for _yearselector in self.yearrangefactor.split(','):
                if '-' in _yearselector:
                    _yearselector = _yearselector.split('-')
                    _yearselectorstart, _yearselectorend = int(_yearselector[0]), int(_yearselector[1])
                    _selectedyears.extend(list(np.arange(_yearselectorstart, _yearselectorend + 1)))
                else:
                    _yearselector = int(_yearselector)
                    _selectedyears.append(_yearselector)

            _selectedyears = sorted(_selectedyears)

            for _yeariterator in self._yearlist[::-1]:
                if _yeariterator not in _selectedyears:
                    _startingcountercolumn = _yearletterdict.get(_yearstartdict.get(_yeariterator)) + _columnstartingcounter
                    sheet.delete_cols(_startingcountercolumn, 12)

        if _selectedyears:
          _length = len(_selectedyears)
          _cellwritingvalue = _length * 12 + 2
        else:
          _cellwritingvalue = _startvalue + 1

        _legendcolorwritingcolumn = 6
        _legendtextwritingcolumn = 6
        for _key in self._colorschemadict:
          sheet[f'{self._findcellvalue(_cellwritingvalue)}{_legendcolorwritingcolumn}'].fill = PatternFill(start_color = self._colorschemadict[_key], fill_type = "solid")
          sheet[f'{self._findcellvalue(_cellwritingvalue + 1)}{_legendtextwritingcolumn}'] = _key
          _legendcolorwritingcolumn = _legendcolorwritingcolumn + 1
          _legendtextwritingcolumn = _legendtextwritingcolumn + 1

        wb.save(self._writingoutputfile)

    def downloadfiles(self):
      self._writingoutputfile

    def main_running_code(self):
        self.load_dataset(file_path,mapping_file_path)
        self.grant_chart_making()
        self.writingdatatoexcel()
        self.downloadfiles()

@app.route('/execute', methods=['POST'])
def execute_function():
    global file_path
    if file_path == '':
        return 'Please upload a file first.'
    try:
        data_analysis = DataAnalysis(file_path, color_range, yearrange, portfolio_type, status_budget_tls, sorting_parameter, ascending)
        result = data_analysis.main_running_code()
        message = 'Output file downloaded successfully!'
        return render_template('index.html', message=message)
    except Exception as e:
        return 'An error occurred: ' + str(e)

@app.route('/del', methods=['POST'])
def del_execute_function():
    global file_path
    if file_path == '':
        return 'Please upload a file first and execute it to clear downloaded file.'
    try:
        def clearcache():
            for file1 in os.listdir('.'):
                if file1 == 'output.xlsx':
                    os.remove(file1)
        result=clearcache()
        message = 'Output File deleted successfully!'
        return render_template('index.html', message=message)
    except Exception as e:
        return 'An error occurred: ' + str(e)

# if __name__ == '__main__':
#     # app.run(debug=True)
#     app.run(host='0.0.0.0', port=5000)
